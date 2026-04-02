"""
Export Dashboard Signals to JSON
Reads the Flatbed Market Signal Dashboard workbooks and exports
market intelligence data for the Pricing Engine to consume.

Run this after updating the dashboard weekly (or whenever LTR data refreshes).
Output: data/dashboard_signals.json

Enriched bridge: now also reads conviction score, cycle phase, regime,
quoting posture, national LTR, and demand signal indicators from the
full dashboard workbook suite.
"""

import csv
import json
import os
import sys
from datetime import datetime

import openpyxl

# --- Paths ---
DASHBOARD_DIR = r"C:\Users\johnb\OneDrive - PS Logistics\Dashboard"
STATE_WB_PATH = os.path.join(DASHBOARD_DIR, "ltr_8day_vs_30day.xlsx")
MARKET_WB_PATH = os.path.join(DASHBOARD_DIR, "market_ltr_8day_vs_30day.xlsx")
REGIME_WB_PATH = os.path.join(DASHBOARD_DIR, "flatbed_ltr_analysis.xlsx")
USMNO_WB_PATH = os.path.join(DASHBOARD_DIR, "Flatbed_IP_USMNO_Enhanced.xlsx")
DEMAND_SIGNALS_DIR = os.path.join(DASHBOARD_DIR, "Dashboard Map Data")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "dashboard_signals.json")

# --- Signal thresholds (for market-level, based on divergence) ---
SIGNAL_THRESHOLDS = [
    (0.60, "Acute Imbalance"),
    (0.50, "Tightening"),
    (0.12, "Firming"),
    (-0.12, "Balanced"),
    (-0.15, "Loosening"),
]
# Below -0.15 = Soft

# --- Pressure score to signal mapping (state-level) ---
PRESSURE_THRESHOLDS = [
    (80, "Acute Imbalance"),
    (65, "Tightening"),
    (45, "Firming"),
    (30, "Balanced"),
    (15, "Loosening"),
]
# Below 15 = Soft

# --- Demand signal indicator definitions (matches build_demand_signals.py) ---
DEMAND_INDICATORS = [
    {'pattern': 'USMNO',  'key': 'ISM_New_Orders',       'label': 'ISM Manufacturing New Orders',      'unit': 'index',           'layers': ['ISM-Grid'],                    'bullish_above': 50},
    {'pattern': 'USTRG',  'key': 'Rig_Count',            'label': 'Baker Hughes Total Rig Count',      'unit': 'rigs',            'layers': ['Energy'],                      'bullish_above': None},
    {'pattern': 'HOUST1F','key': 'Housing_Starts',        'label': 'Housing Starts (SF, SAAR)',         'unit': 'thousands',       'layers': ['Permits-Starts','Construction'],'bullish_above': None},
    {'pattern': 'PERMIT1','key': 'Building_Permits',      'label': 'Building Permits (SF)',             'unit': 'thousands',       'layers': ['Permits-Starts','Construction'],'bullish_above': None},
    {'pattern': 'TLMFGCON','key':'Construction_Spending',  'label': 'Construction Spending (Mfg)',       'unit': 'millions_USD',    'layers': ['Construction'],                'bullish_above': None},
    {'pattern': 'LBR1',   'key': 'Lumber_Futures',        'label': 'Lumber Futures',                    'unit': 'USD_per_mbf',     'layers': ['Lumber','Lumber-PNW'],          'bullish_above': None},
    {'pattern': 'HRC1',   'key': 'HRC_Steel',             'label': 'Hot-Rolled Coil Steel Futures',     'unit': 'USD_per_ton',     'layers': ['Steel'],                       'bullish_above': None},
    {'pattern': 'CL1',    'key': 'Crude_Oil_WTI',         'label': 'Crude Oil (WTI) Futures',           'unit': 'USD_per_barrel',  'layers': ['Energy'],                      'bullish_above': None},
    {'pattern': 'ZC1',    'key': 'Corn_Futures',           'label': 'Corn Futures',                      'unit': 'cents_per_bushel','layers': ['HeavyEquip-Ag'],               'bullish_above': None},
    {'pattern': 'ZS1',    'key': 'Soybean_Futures',        'label': 'Soybean Futures',                   'unit': 'cents_per_bushel','layers': ['HeavyEquip-Ag'],               'bullish_above': None},
    {'pattern': 'ZW1',    'key': 'Wheat_Futures',          'label': 'Wheat Futures',                     'unit': 'cents_per_bushel','layers': ['HeavyEquip-Ag'],               'bullish_above': None},
    {'pattern': 'FAN',    'key': 'Wind_Energy_ETF',        'label': 'First Trust Global Wind Energy ETF','unit': 'USD',             'layers': ['Renewables'],                  'bullish_above': None},
]


def classify_signal_from_divergence(pct_diff):
    """Classify market signal from 8D vs 30D divergence percentage."""
    if pct_diff is None:
        return "Balanced"
    for threshold, signal in SIGNAL_THRESHOLDS:
        if pct_diff >= threshold:
            return signal
    return "Soft"


def divergence_to_pressure_score(pct_diff, signal):
    """Estimate a pressure score from divergence for market-level data.
    State-level has actual computed scores; market-level only has divergence.
    We estimate using the midpoint of each signal's pressure range."""
    score_map = {
        "Acute Imbalance": 90,
        "Tightening": 72,
        "Firming": 55,
        "Balanced": 37,
        "Loosening": 22,
        "Soft": 7,
    }
    base = score_map.get(signal, 37)
    # Refine within the band using the divergence magnitude
    if signal == "Firming" and pct_diff is not None:
        ratio = min(max((pct_diff - 0.12) / 0.38, 0), 1)
        base = 45 + ratio * 19
    elif signal == "Tightening" and pct_diff is not None:
        ratio = min(max((pct_diff - 0.50) / 0.10, 0), 1)
        base = 65 + ratio * 14
    elif signal == "Acute Imbalance" and pct_diff is not None:
        ratio = min(max((pct_diff - 0.60) / 0.40, 0), 1)
        base = 80 + ratio * 20
    elif signal == "Balanced" and pct_diff is not None:
        ratio = min(max((pct_diff + 0.12) / 0.24, 0), 1)
        base = 30 + ratio * 14
    elif signal == "Loosening" and pct_diff is not None:
        ratio = min(max((pct_diff + 0.15) / 0.03, 0), 1)
        base = 15 + ratio * 14
    elif signal == "Soft" and pct_diff is not None:
        ratio = min(max((pct_diff + 0.50) / 0.35, 0), 1)
        base = ratio * 14
    return round(base, 1)


def map_market_name_to_dat_code(market_name, state_abbrev):
    """Map dashboard market names like 'Birmingham, AL' to DAT codes like 'AL_BIR'.
    Uses first 3 letters of city + state abbreviation."""
    if not market_name or not state_abbrev:
        return None
    city = market_name.split(",")[0].strip()
    city_code = city[:3].upper()
    return f"{state_abbrev}_{city_code}"


# ──────────────────────────────────────────────────────────────────────────────
# NEW: Conviction, Phase, Regime readers
# ──────────────────────────────────────────────────────────────────────────────

def read_regime_dashboard(wb_path):
    """Read conviction score, output label, operational bias, and national
    LTR from the Regime_Dashboard and Data_Weekly sheets of
    flatbed_ltr_analysis.xlsx.

    Returns dict with:
      conviction_score, monthly_structural, weekly_momentum,
      output_label, operational_bias, summary,
      national_ltr, national_ema, national_roc,
      ltr_direction (RISING/FALLING based on 4-week RoC)
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)

        # --- Regime_Dashboard sheet ---
        ws = wb["Regime_Dashboard"]
        result["monthly_structural_score"] = _safe_num(ws.cell(5, 2).value)
        result["weekly_momentum_score"] = _safe_num(ws.cell(6, 2).value)
        result["conviction_score"] = _safe_num(ws.cell(7, 2).value)
        result["output_label"] = _safe_str(ws.cell(9, 2).value)
        result["operational_bias"] = _safe_str(ws.cell(10, 2).value)
        result["summary"] = _safe_str(ws.cell(12, 2).value)
        result["monthly_date"] = _safe_date(ws.cell(3, 2).value)
        result["weekly_date"] = _safe_date(ws.cell(3, 5).value)

        # --- Data_Weekly sheet (last row for national LTR) ---
        ws2 = wb["Data_Weekly"]
        max_r = ws2.max_row
        # Walk backwards to find last non-empty row
        for r in range(max_r, 1, -1):
            ltr_val = ws2.cell(r, 4).value  # Col D = National Flatbed LTR
            if ltr_val is not None:
                result["national_ltr"] = round(float(ltr_val), 2)
                result["national_ema"] = _safe_round(ws2.cell(r, 5).value, 2)  # Col E = 8 Period EMA
                result["national_roc"] = _safe_round(ws2.cell(r, 6).value, 4)  # Col F = 4 Week RoC
                week_date = ws2.cell(r, 1).value  # Col A = Week Ending
                result["national_ltr_date"] = _safe_date(week_date)
                # Derive direction from RoC
                roc = result["national_roc"]
                if roc is not None:
                    result["ltr_direction"] = "RISING" if roc > 0 else "FALLING"
                else:
                    result["ltr_direction"] = "UNKNOWN"
                break

        wb.close()
    except Exception as e:
        print(f"WARNING: Could not read regime dashboard: {e}")
    return result


def read_regime_from_usmno(wb_path):
    """Read the macro regime from the last row of the USMNO Integration sheet
    in Flatbed_IP_USMNO_Enhanced.xlsx.

    Returns dict with: regime (e.g. 'EXPANSION'), regime_date
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb["USMNO Integration"]
        max_r = ws.max_row
        for r in range(max_r, 1, -1):
            date_val = ws.cell(r, 1).value
            if date_val is None:
                continue
            # Column 21 = Final Regime (primary), Column 7/12 = Option A/C fallback
            regime_final = _safe_str(ws.cell(r, 21).value)
            regime_a = _safe_str(ws.cell(r, 7).value)
            regime_c = _safe_str(ws.cell(r, 12).value)
            regime = regime_final or regime_c or regime_a
            if regime:
                result["regime"] = regime
                result["regime_date"] = _safe_date(date_val)
                break
        wb.close()
    except Exception as e:
        print(f"WARNING: Could not read USMNO regime: {e}")
    return result


def _safe_num(val):
    """Convert to number or None."""
    if val is None:
        return None
    try:
        return round(float(val), 4)
    except (ValueError, TypeError):
        return None


def _safe_round(val, decimals=2):
    """Round a numeric value or return None."""
    if val is None:
        return None
    try:
        return round(float(val), decimals)
    except (ValueError, TypeError):
        return None


def _safe_str(val):
    """Convert to clean string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _safe_date(val):
    """Convert datetime to ISO date string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


# ──────────────────────────────────────────────────────────────────────────────
# NEW: Demand signal indicator reader
# ──────────────────────────────────────────────────────────────────────────────

def parse_date_str(s):
    """Auto-detect date format from CSV."""
    s = s.strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%dT%H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def find_indicator_csv(pattern):
    """Find the CSV file matching the indicator pattern in the Dashboard Map Data folder."""
    if not os.path.exists(DEMAND_SIGNALS_DIR):
        return None
    for f in os.listdir(DEMAND_SIGNALS_DIR):
        if not f.lower().endswith('.csv'):
            continue
        cleaned = f.replace('!', '').replace(',', '').replace(' ', '')
        if pattern in cleaned or pattern in f:
            return os.path.join(DEMAND_SIGNALS_DIR, f)
    return None


def read_demand_indicators():
    """Read the last 6 months of each demand signal indicator from CSVs.
    Returns a dict of indicator_key -> {label, unit, layers, latest, prior, direction, data[]}
    """
    indicators = {}
    for ind in DEMAND_INDICATORS:
        fp = find_indicator_csv(ind['pattern'])
        if not fp:
            continue

        rows = []
        try:
            with open(fp, 'r', encoding='utf-8-sig') as fh:
                reader = csv.reader(fh)
                next(reader)  # skip header
                for row in reader:
                    if len(row) < 2 or not row[0].strip() or not row[1].strip():
                        continue
                    dt = parse_date_str(row[0])
                    if dt is None:
                        continue
                    try:
                        val = float(row[1].replace(',', ''))
                    except ValueError:
                        continue
                    rows.append((dt, val))
            rows.sort(key=lambda x: x[0])
        except Exception as e:
            print(f"  WARNING: Could not read {ind['key']}: {e}")
            continue

        if len(rows) < 2:
            continue

        # Last 6 months for the bridge (keep it compact)
        recent = rows[-6:]
        latest_val = recent[-1][1]
        prior_val = recent[-2][1]

        # Direction
        if latest_val > prior_val * 1.005:
            direction = "RISING"
        elif latest_val < prior_val * 0.995:
            direction = "FALLING"
        else:
            direction = "FLAT"

        # 3-month-ago for trend context
        three_mo_ago = rows[-4][1] if len(rows) >= 4 else None

        indicators[ind['key']] = {
            "label": ind['label'],
            "unit": ind['unit'],
            "layers": ind['layers'],
            "latest": round(latest_val, 2),
            "prior": round(prior_val, 2),
            "three_mo_ago": round(three_mo_ago, 2) if three_mo_ago else None,
            "direction": direction,
            "latest_date": recent[-1][0].strftime("%Y-%m"),
            "bullish_above": ind.get('bullish_above'),
            "data": [{"d": dt.strftime("%Y-%m"), "v": round(v, 2)} for dt, v in recent],
        }

    return indicators


# ──────────────────────────────────────────────────────────────────────────────
# Quoting posture logic (derives from conviction + regime)
# ──────────────────────────────────────────────────────────────────────────────

def derive_quoting_posture(conviction_score, regime, ltr_direction):
    """Derive a quoting posture recommendation from conviction + regime + direction.

    Returns dict with posture label + margin guidance.
    """
    if conviction_score is None:
        return {"posture": "NEUTRAL", "margin_bias": 0.0, "guidance": "Insufficient data for posture"}

    cs = int(conviction_score)

    # Conviction score ranges -6 to +6
    # Positive = tightening pressure, negative = loosening pressure
    if cs >= 4:
        posture = "DEFENSIVE"
        margin_bias = 0.03  # +3% margin floor
        guidance = "Market tightening strongly. Protect cost risk, quote conservatively."
    elif cs >= 2:
        posture = "CAUTIOUS"
        margin_bias = 0.015  # +1.5%
        guidance = "Market firming. Lean toward higher margins, limit aggressive discounting."
    elif cs >= -1:
        posture = "NEUTRAL"
        margin_bias = 0.0
        guidance = "Market balanced. Standard margin targets apply."
    elif cs >= -3:
        posture = "OPPORTUNISTIC"
        margin_bias = -0.01  # -1% — can be slightly more aggressive
        guidance = "Market softening. Opportunity to win volume with competitive pricing."
    else:
        posture = "AGGRESSIVE"
        margin_bias = -0.02  # -2%
        guidance = "Market in oversupply. Pursue volume aggressively."

    # If regime contradicts conviction, note the tension
    tension = None
    if regime == "CONTRACTION" and cs >= 3:
        tension = "High conviction tightening during contraction — likely transitional; monitor weekly."
    elif regime == "EXPANSION" and cs <= -3:
        tension = "Deep loosening during expansion — unusual; verify demand signal data."

    return {
        "posture": posture,
        "margin_bias": margin_bias,
        "guidance": guidance,
        "tension": tension,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Cycle phase mapping (for engine.py get_cycle_buffer compatibility)
# ──────────────────────────────────────────────────────────────────────────────

def derive_phase(conviction_score, ltr_direction, regime):
    """Map conviction + direction + regime to a numeric phase (0-5) that
    engine.py's get_cycle_buffer() understands.

    Phase mapping:
      0 = Bottom / Recovery start (low conviction, direction turning up)
      1 = Early expansion (moderate conviction rising)
      2 = Mid expansion (stable, moderate)
      3 = Late expansion / peak (high conviction, rising slowing)
      4 = Early contraction (conviction turning negative)
      5 = Deep contraction (strong negative conviction)
    """
    if conviction_score is None:
        return 2  # Default to mid-expansion (safest neutral)

    cs = int(conviction_score)
    rising = (ltr_direction == "RISING")

    if regime == "EXPANSION":
        if cs <= 0:
            return 1 if rising else 0  # Early expansion or bottom
        elif cs <= 2:
            return 2  # Mid expansion
        elif cs <= 4:
            return 3 if not rising else 4  # Late expansion or early contraction signal
        else:
            return 4  # Peak / transition
    else:  # CONTRACTION or unknown
        if cs >= 3:
            return 4  # Early contraction with tightening (rebound signal)
        elif cs >= 0:
            return 3  # Late contraction, stabilizing
        else:
            return 5 if cs <= -3 else 4  # Deep or early contraction


# ══════════════════════════════════════════════════════════════════════════════
# MAIN EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def export_signals():
    """Main export function. Reads dashboard workbooks and writes JSON."""
    print(f"Reading state workbook: {STATE_WB_PATH}")
    state_wb = openpyxl.load_workbook(STATE_WB_PATH, data_only=True)

    # --- Extract state-level current data ---
    current_sheet = state_wb["Flatbed 8d vs 30d "] if "Flatbed 8d vs 30d " in state_wb.sheetnames else state_wb["Flatbed 8d vs 30d"]
    states = {}
    as_of_date = None
    for row in current_sheet.iter_rows(min_row=2, max_col=5, values_only=True):
        state, ltr_8d, ltr_30d, date_val, pct_diff = row
        if state is None or ltr_8d is None:
            continue
        state = str(state).strip()
        if len(state) != 2:
            continue
        if date_val and as_of_date is None:
            if isinstance(date_val, datetime):
                as_of_date = date_val.strftime("%Y-%m-%d")
            else:
                as_of_date = str(date_val)
        states[state] = {
            "ltr_8d": round(float(ltr_8d), 1) if ltr_8d else None,
            "ltr_30d": round(float(ltr_30d), 1) if ltr_30d else None,
            "divergence": round(float(pct_diff), 4) if pct_diff else None,
        }

    # --- Extract state-level pressure scores from history (most recent week) ---
    history_sheet = state_wb["LTR_State_History"]
    headers = [cell.value for cell in next(history_sheet.iter_rows(min_row=1, max_row=1))]

    # Find column indices
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    # Read all rows, keep only the most recent date per state
    latest_data = {}
    latest_date = None
    national_data = {}
    for row in history_sheet.iter_rows(min_row=2, values_only=True):
        values = list(row)
        date_val = values[col_map.get("Report_Date", 0)]
        state = values[col_map.get("State", 1)]
        if date_val is None or state is None:
            continue

        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)

        if latest_date is None or date_str >= latest_date:
            if date_str > (latest_date or ""):
                latest_date = date_str
                latest_data = {}
            state = str(state).strip()

            pressure_score = values[col_map.get("Pressure_Score", 14)]
            pressure_state = values[col_map.get("Pressure_State", 15)]
            momentum = values[col_map.get("Momentum", 20)]
            pct_diff_zscore = values[col_map.get("PctDiff_ZScore", 18)]

            entry = {
                "pressure_score": round(float(pressure_score), 1) if pressure_score is not None else None,
                "signal": str(pressure_state) if pressure_state else None,
                "momentum": round(float(momentum), 4) if momentum is not None else None,
                "z_score": round(float(pct_diff_zscore), 2) if pct_diff_zscore is not None else None,
            }
            latest_data[state] = entry

            # Extract national data from any row
            nat_8d = values[col_map.get("National_LTR_8D", 6)]
            nat_30d = values[col_map.get("National_LTR_30D", 7)]
            nat_pct = values[col_map.get("National_Pct_Diff", 9)]
            if nat_8d is not None:
                national_data = {
                    "ltr_8d": round(float(nat_8d), 1),
                    "ltr_30d": round(float(nat_30d), 1) if nat_30d else None,
                    "divergence": round(float(nat_pct), 4) if nat_pct else None,
                }

    # Merge pressure data into state entries
    for state, pressure in latest_data.items():
        if state in states:
            states[state].update(pressure)
        else:
            states[state] = pressure

    state_wb.close()

    # --- Extract market-level data ---
    print(f"Reading market workbook: {MARKET_WB_PATH}")
    market_wb = openpyxl.load_workbook(MARKET_WB_PATH, data_only=True)
    market_sheet = market_wb["Market_Current_Week"]

    markets = {}
    for row in market_sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        state_name, state_abbrev, market_name, ltr_8d, ltr_30d, pct_diff, div_signal = row
        if market_name is None or state_abbrev is None:
            continue

        state_abbrev = str(state_abbrev).strip()
        market_name = str(market_name).strip()

        # Map to DAT code
        dat_code = map_market_name_to_dat_code(market_name, state_abbrev)
        if dat_code is None:
            continue

        signal = str(div_signal) if div_signal else classify_signal_from_divergence(pct_diff)
        est_pressure = divergence_to_pressure_score(float(pct_diff) if pct_diff else None, signal)

        markets[dat_code] = {
            "market_name": market_name,
            "state": state_abbrev,
            "ltr_8d": round(float(ltr_8d), 1) if ltr_8d else None,
            "ltr_30d": round(float(ltr_30d), 1) if ltr_30d else None,
            "divergence": round(float(pct_diff), 4) if pct_diff else None,
            "signal": signal,
            "pressure_score": est_pressure,
        }

    market_wb.close()

    # ──────────────────────────────────────────────────────────────────────
    # NEW: Read conviction, regime, and demand indicators
    # ──────────────────────────────────────────────────────────────────────
    print(f"Reading regime dashboard: {REGIME_WB_PATH}")
    regime_data = read_regime_dashboard(REGIME_WB_PATH)

    print(f"Reading USMNO regime: {USMNO_WB_PATH}")
    usmno_data = read_regime_from_usmno(USMNO_WB_PATH)

    print(f"Reading demand signal indicators from: {DEMAND_SIGNALS_DIR}")
    demand_indicators = read_demand_indicators()

    # Merge national data with regime data
    national_ltr_from_regime = regime_data.get("national_ltr")
    if national_ltr_from_regime is not None:
        national_data["national_ltr_weekly"] = national_ltr_from_regime
        national_data["national_ema"] = regime_data.get("national_ema")
        national_data["national_roc"] = regime_data.get("national_roc")
        national_data["ltr_direction"] = regime_data.get("ltr_direction")

    # Derive quoting posture
    conviction = regime_data.get("conviction_score")
    regime = usmno_data.get("regime", "UNKNOWN")
    ltr_dir = regime_data.get("ltr_direction", "UNKNOWN")
    posture_data = derive_quoting_posture(conviction, regime, ltr_dir)

    # Derive engine-compatible phase
    phase = derive_phase(conviction, ltr_dir, regime)

    # --- Build the output JSON ---
    output = {
        "exported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "as_of": as_of_date or latest_date,

        # National-level (enriched)
        "national": national_data,

        # Macro intelligence (NEW)
        "macro": {
            "regime": regime,
            "regime_date": usmno_data.get("regime_date"),
            "phase": phase,
            "conviction_score": int(conviction) if conviction is not None else None,
            "monthly_structural_score": regime_data.get("monthly_structural_score"),
            "weekly_momentum_score": regime_data.get("weekly_momentum_score"),
            "output_label": regime_data.get("output_label"),
            "operational_bias": regime_data.get("operational_bias"),
            "summary": regime_data.get("summary"),
        },

        # Quoting posture (NEW)
        "quoting_posture": posture_data,

        # State and market data (existing)
        "states": states,
        "markets": markets,

        # Demand signal indicators (NEW)
        "demand_indicators": demand_indicators,

        # Signal thresholds (existing)
        "signal_thresholds": {
            "Acute Imbalance": {"min_pressure": 80, "min_divergence": 0.60},
            "Tightening": {"min_pressure": 65, "min_divergence": 0.50},
            "Firming": {"min_pressure": 45, "min_divergence": 0.12},
            "Balanced": {"min_pressure": 30, "min_divergence": -0.12},
            "Loosening": {"min_pressure": 15, "min_divergence": -0.15},
            "Soft": {"min_pressure": 0, "min_divergence": None},
        },
    }

    # Write JSON
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*60}")
    print(f"Exported to: {OUTPUT_PATH}")
    print(f"As of: {output['as_of']}")
    print(f"{'='*60}")
    print(f"\n--- NATIONAL ---")
    print(f"  National LTR (weekly): {national_data.get('national_ltr_weekly', 'N/A')}")
    print(f"  Direction: {national_data.get('ltr_direction', 'N/A')}")
    print(f"  EMA: {national_data.get('national_ema', 'N/A')}")
    print(f"  RoC: {national_data.get('national_roc', 'N/A')}")

    print(f"\n--- MACRO ---")
    macro = output["macro"]
    print(f"  Regime: {macro['regime']}")
    print(f"  Phase: {macro['phase']}")
    print(f"  Conviction: {macro['conviction_score']}")
    print(f"  Label: {macro['output_label']}")
    print(f"  Bias: {macro['operational_bias']}")

    print(f"\n--- QUOTING POSTURE ---")
    qp = output["quoting_posture"]
    print(f"  Posture: {qp['posture']}")
    print(f"  Margin Bias: {qp['margin_bias']:+.1%}")
    print(f"  Guidance: {qp['guidance']}")
    if qp.get("tension"):
        print(f"  Tension: {qp['tension']}")

    print(f"\n--- COVERAGE ---")
    print(f"  States: {len(output['states'])}")
    print(f"  Markets: {len(output['markets'])}")
    print(f"  Demand Indicators: {len(output['demand_indicators'])}")

    # State signal distribution
    signal_counts = {}
    for s_data in output["states"].values():
        sig = s_data.get("signal", "Unknown")
        signal_counts[sig] = signal_counts.get(sig, 0) + 1
    print(f"\n--- STATE SIGNAL DISTRIBUTION ---")
    for sig, count in sorted(signal_counts.items()):
        print(f"  {sig}: {count}")

    # Demand indicator summary
    if demand_indicators:
        print(f"\n--- DEMAND INDICATORS ---")
        rising = sum(1 for v in demand_indicators.values() if v["direction"] == "RISING")
        falling = sum(1 for v in demand_indicators.values() if v["direction"] == "FALLING")
        flat = sum(1 for v in demand_indicators.values() if v["direction"] == "FLAT")
        print(f"  Rising: {rising}  |  Falling: {falling}  |  Flat: {flat}")
        for k, v in demand_indicators.items():
            arrow = {"RISING": "+", "FALLING": "-", "FLAT": "="}[v["direction"]]
            print(f"    [{arrow}] {v['label']}: {v['latest']} {v['unit']}")

    return output


if __name__ == "__main__":
    export_signals()
