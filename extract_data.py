import json
import os
import pandas as pd

WORKBOOK = os.path.join(os.path.dirname(__file__), "Master_Flatbed_RFQ_Pricing_Engine.xlsx")
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")


def extract_matrix(sheet_name, output_name):
    df = pd.read_excel(WORKBOOK, sheet_name=sheet_name, index_col=0, engine="openpyxl")
    df.to_csv(os.path.join(DATA_DIR, output_name))
    print(f"  {sheet_name} -> {output_name} ({df.shape[0]} x {df.shape[1]})")


def extract_zip3():
    df = pd.read_excel(WORKBOOK, sheet_name="Zip3 Lookup", engine="openpyxl", dtype={"Zip3": str})
    df.columns = ["Zip3", "State", "DAT_Market", "Market_City", "Distance"]
    df["Zip3"] = df["Zip3"].apply(lambda x: str(x).zfill(3))
    df.to_csv(os.path.join(DATA_DIR, "zip3_lookup.csv"), index=False)
    print(f"  Zip3 Lookup -> zip3_lookup.csv ({len(df)} rows)")


def extract_city():
    df = pd.read_excel(WORKBOOK, sheet_name="City Lookup", engine="openpyxl")
    df.columns = ["City", "State", "DAT_Market"]
    df.to_csv(os.path.join(DATA_DIR, "city_lookup.csv"), index=False)
    print(f"  City Lookup -> city_lookup.csv ({len(df)} rows)")


def extract_control_panel():
    import openpyxl
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["Control Panel"]
    vol_6 = {}
    for tier, row in [("LOW", 34), ("MED", 35), ("HIGH", 36)]:
        vol_6[tier] = {}
        for col, p in [(4, "50"), (5, "55"), (6, "60"), (7, "65"), (8, "70"), (9, "75"), (10, "80"), (11, "85"), (12, "90")]:
            vol_6[tier][p] = ws.cell(row, col).value
    vol_12 = {}
    for tier, row in [("LOW", 40), ("MED", 41), ("HIGH", 42)]:
        vol_12[tier] = {}
        for col, p in [(4, "50"), (5, "55"), (6, "60"), (7, "65"), (8, "70"), (9, "75"), (10, "80"), (11, "85"), (12, "90")]:
            vol_12[tier][p] = ws.cell(row, col).value

    params = {
        "regime": ws.cell(7, 4).value,
        "phase": ws.cell(8, 4).value,
        "ltr_direction": ws.cell(9, 4).value,
        "national_ltr": ws.cell(10, 4).value,
        "ltr_smooth": ws.cell(11, 4).value,
        "target_margin": ws.cell(15, 4).value,
        "contract_term": ws.cell(16, 4).value,
        "confidence": ws.cell(17, 4).value,
        "fsc_per_mile": ws.cell(18, 4).value,
        "green_zone": ws.cell(21, 4).value,
        "red_zone": ws.cell(22, 4).value,
        "gate_tolerance": ws.cell(23, 4).value,
        "dat_file_date": str(ws.cell(53, 4).value),
        "vol_buffer_table": {"6": vol_6, "12": vol_12},
        "liq_adj_table": {"THIN": 0.03, "MODERATE": 0.01, "GOOD": 0.0, "DEEP": -0.01},
    }
    wb.close()
    with open(os.path.join(DATA_DIR, "control_panel.json"), "w") as f:
        json.dump(params, f, indent=2, default=str)
    print(f"  Control Panel -> control_panel.json")


if __name__ == "__main__":
    os.makedirs(DATA_DIR, exist_ok=True)
    print("Extracting matrices (this may take a minute)...")
    for sheet, fname in [("Rate Matrix", "rate_matrix.csv"), ("StdDev Matrix", "stddev_matrix.csv"),
                         ("Reports Matrix", "reports_matrix.csv"), ("Miles Matrix", "miles_matrix.csv")]:
        extract_matrix(sheet, fname)
    print("Extracting lookups...")
    extract_zip3()
    extract_city()
    print("Extracting control panel...")
    extract_control_panel()
    print("Done. Data files are in:", DATA_DIR)
